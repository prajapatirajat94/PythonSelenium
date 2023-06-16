from openpyxl import load_workbook
import openpyxl


Workbook=openpyxl.load_workbook("myexcel.xlsx")

Sheet=Workbook.active

numrow=Sheet.max_row
numcol=Sheet.max_column

print("number of row is : ",numrow)
print("number of column is : ",numcol)

for i in range(1,numrow+1):
    for k in range(1,numcol+1):
        print(Sheet.cell(row=i,column=k).value)

print("--------------------------**********-------------------------")
for j in range(1,numcol+1):
    print(Sheet.cell(row=1,column=j).value)
print("--------------------------**********-------------------------")
table_value=Sheet['A1':'B4']
MyValue={}
for cell1,cell2 in table_value:
    #MyValue.update(cell1.value : cell2.value)
    data={cell1.value:cell2.value}
 #   MyValue.update({"Rajat": "Kamal"})
    MyValue.update(data)
    print(cell1.value,cell2.value)

for i in  MyValue:
    print("this value is comig from dict: "+MyValue[i])



# Python program to read an excel file

